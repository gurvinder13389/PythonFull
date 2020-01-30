import openpyxl
import os
from MTAF.ReusableFunctions import globalVar
from openpyxl.styles import PatternFill


def getProjectPath():
    if not globalVar.bln_SkipTestCase:
        try:
            return os.path.join(os.path.dirname(__file__), "..")
        except Exception as ex:
            print('Project location cannot be fetched. ' + str(ex))
            globalVar.bln_SkipTestCaseLog = True
            globalVar.bln_SkipTestCase = True
            return None


def getExcelWB(readmode = True):
    if not globalVar.bln_SkipTestCase:
        try:
            if readmode:
                return openpyxl.load_workbook(filename = getProjectPath() + "/TestData/" + globalVar.excelWBName + ".xlsx",read_only=True)
            else:
                return openpyxl.load_workbook(filename = getProjectPath() + "/TestData/" + globalVar.excelWBName + ".xlsx")
        except Exception as ex:
            print('File ' + getProjectPath() +  "/TestData/" + globalVar.excelWBName + ' cannot be loaded. ' + str(ex))
            globalVar.bln_SkipTestCaseLog = True
            globalVar.bln_SkipTestCase = True
            return None


def getRowCount(sheetName):
    if not globalVar.bln_SkipTestCase:
        try:
            return globalVar.excelWB[sheetName].max_row
        except Exception as ex:
            print('Row count cannot be fetched for excel file ' + globalVar.excelWBName + ' for sheet ' + sheetName + ". " + str(ex))
            globalVar.bln_SkipTestCaseLog = True
            globalVar.bln_SkipTestCase = True
            return None


def getColumnCount(sheetName):
    if not globalVar.bln_SkipTestCase:
        try:
            return globalVar.excelWB[sheetName].max_column
        except Exception as ex:
            print('Column count cannot be fetched for excel file ' + globalVar.excelWBName + ' for sheet ' + sheetName + ". " + str(ex))
            globalVar.bln_SkipTestCaseLog = True
            globalVar.bln_SkipTestCase = True
            return None


def getData(sheetName,rowNum,columnNum):
    if not globalVar.bln_SkipTestCase:
        try:
            if globalVar.excelWB[sheetName].cell(row=rowNum, column=columnNum).value == None:
                return ''
            else:
                return str(globalVar.excelWB[sheetName].cell(row=rowNum, column=columnNum).value)
        except Exception as ex:
            print('Data for cell with row number ' + str(rowNum) + ' and column number ' + str(columnNum) + ' cannot be fetched for excel file ' + globalVar.excelWBName + ' for sheet ' + sheetName + ". " + str(ex))
            globalVar.bln_SkipTestCaseLog = True
            globalVar.bln_SkipTestCase = True
            return None


def findRowPos(sheetName,searchValue,colNum):
    if not globalVar.bln_SkipTestCase:
        try:
            ws = globalVar.excelWB[sheetName]
            totalRowCount = ws.max_row
            blnStatus = True
            rowNum = 1
            while blnStatus:
                tempData = ws.cell(row=rowNum, column=colNum).value
                if rowNum > totalRowCount:
                    rowNum = 0
                    blnStatus = False
                elif searchValue == tempData:
                    blnStatus = False
                else:
                    rowNum+=1
            return rowNum
        except Exception as ex:
            print('Row position cannot be fetched for excel file ' + globalVar.excelWBName + ' for sheet ' + sheetName + ' at column ' + str(colNum) + ". " + str(ex))
            globalVar.bln_SkipTestCaseLog = True
            globalVar.bln_SkipTestCase = True
            return None


def findColumnPos(sheetName,searchValue,rowNum):
    if not globalVar.bln_SkipTestCase:
        try:
            ws = globalVar.excelWB.get_sheet_by_name(sheetName)
            totalColumnCount = ws.max_column
            blnStatus = True
            colNum = 1
            while blnStatus:
                tempData = ws.cell(row=rowNum, column=colNum)
                if colNum > totalColumnCount:
                    colNum = 0
                    blnStatus = False
                elif searchValue == tempData:
                    blnStatus = False
                else:
                    colNum += 1
            return colNum
        except Exception as ex:
            print('Column position cannot be fetched for excel file ' + globalVar.excelWBName + ' for sheet ' + sheetName + ' at row ' + str(rowNum) + ". " + str(ex))
            globalVar.bln_SkipTestCaseLog = True
            globalVar.bln_SkipTestCase = True
            return None


def getTestCaseData(sheetName):
    if not globalVar.bln_SkipTestCase:
        try:
            globalVar.excelWBName = globalVar.moduleName
            globalVar.excelWB = getExcelWB(False)
            iRowCount = getRowCount(sheetName)
            iColCount = getColumnCount(sheetName)
            dictData = dict()
            for iCol in range(1, int(iColCount)+1):
                strColName = str(getData(sheetName,1,iCol))
                dictData[strColName] = dict()
                for iRow in range(2, iRowCount + 1):
                    strMode = getData(sheetName,iRow,2)
                    if strMode == None:
                        strMode = ''
                    if strMode.lower() != 'n':
                        dictData[strColName][iRow] = getData(sheetName,iRow,iCol)
            return dictData
        except Exception as ex:
            print(
                'Test Data cannot be fetched for excel file ' + globalVar.excelWBName + ' for sheet ' + sheetName + ". " + str(ex))
            globalVar.bln_SkipTestCaseLog = True
            globalVar.bln_SkipTestCase = True
            return None


def saveExcelWB(filePath):
    if not globalVar.bln_SkipTestCase:
        try:
            globalVar.excelWB.save(filename = filePath)
        except Exception as ex:
            print('Excel workbook at the path ' + filePath + ' cannot be saved. ' + str(ex))
            globalVar.bln_SkipTestCaseLog = True
            globalVar.bln_SkipTestCase = True


def setCellData(sheetName,rowNum,columnNum,strData):
    if not globalVar.bln_SkipTestCase:
        try:
            globalVar.excelWB[sheetName].cell(row=rowNum, column=columnNum).value = strData
            saveExcelWB(getProjectPath() + "/TestData/" + globalVar.excelWBName + ".xlsx")
        except Exception as ex:
            print('Data on cell with row number ' + str(rowNum) + ' and column number ' + str(columnNum) + ' cannot be set for excel file ' + globalVar.excelWBName + ' for sheet ' + sheetName + ". " + str(ex))
            globalVar.bln_SkipTestCaseLog = True
            globalVar.bln_SkipTestCase = True


def fillCellColor(sheetName,rowNum,columnNum,strColor):
    if not globalVar.bln_SkipTestCase:
        try:
            strFill = PatternFill(start_color=strColor, end_color=strColor,fill_type='solid')
            globalVar.excelWB[sheetName].cell(row=rowNum, column=columnNum).fill = strFill
            saveExcelWB(getProjectPath() + "/TestData/" + globalVar.excelWBName + ".xlsx")
        except Exception as ex:
            print('Cell with row number ' + str(rowNum) + ' and column number ' + str(columnNum) + ' cannot be filled with color ' + strColor + ' for excel file ' + globalVar.excelWBName + ' for sheet ' + sheetName + ". " + str(ex))
            globalVar.bln_SkipTestCaseLog = True
            globalVar.bln_SkipTestCase = True


def removeCellColor(sheetName,rowNum,columnNum):
    if not globalVar.bln_SkipTestCase:
        try:
            strFill = PatternFill(fill_type=None)
            globalVar.excelWB[sheetName].cell(row=rowNum, column=columnNum).fill = strFill
            saveExcelWB(getProjectPath() + "/TestData/" + globalVar.excelWBName + ".xlsx")
        except Exception as ex:
            print('Color on Cell with row number ' + str(rowNum) + ' and column number ' + str(columnNum) + ' cannot be removed for excel file ' + globalVar.excelWBName + ' for sheet ' + sheetName + ". " + str(ex))
            globalVar.bln_SkipTestCaseLog = True
            globalVar.bln_SkipTestCase = True
