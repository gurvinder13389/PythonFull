import json
import openpyxl
import os

jsonObj = open(os.path.dirname(__file__) + "/TestData.JSON")
JsonData = json.loads(jsonObj.read())
jsonList = JsonData["data"]["items"]
jsonObj.close()

excelWB = openpyxl.load_workbook(filename = os.path.dirname(__file__) + "/TestData.xlsx")

iRowStart = excelWB["Sheet1"].max_row + 1

for itm in jsonList:
    print(itm["name"])
    excelWB["Sheet1"].cell(row=iRowStart, column=1).value = itm["name"]
    excelWB["Sheet1"].cell(row=iRowStart, column=2).value = itm["category"]["name"]
    excelWB["Sheet1"].cell(row=iRowStart, column=3).value = itm["provider"]["logo"]
    strLoc = ""
    for loc in itm["provider"]["locations"]:
        strLoc = strLoc + "," + loc["name"]
    strLoc = strLoc[1:]
    excelWB["Sheet1"].cell(row=iRowStart, column=4).value = strLoc
    excelWB["Sheet1"].cell(row=iRowStart, column=5).value = itm["educationLevelId"]
    excelWB["Sheet1"].cell(row=iRowStart, column=6).value = itm["provider"]["name"]
    iRowStart = iRowStart + 1
excelWB.save(filename = "C:/Users/gurvinder122724/Desktop/JSONPython/TestData.xlsx")
